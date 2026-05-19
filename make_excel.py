from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "NcConnect Friction Report"

header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F3542", end_color="2F3542", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

crit_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
crit_font = Font(name="Calibri", bold=True, size=10, color="CC0000")
high_fill = PatternFill(start_color="FFF0DD", end_color="FFF0DD", fill_type="solid")
high_font = Font(name="Calibri", bold=True, size=10, color="CC6600")
med_fill = PatternFill(start_color="FFFCE0", end_color="FFFCE0", fill_type="solid")
med_font = Font(name="Calibri", bold=True, size=10, color="997A00")
low_fill = PatternFill(start_color="E0F5EC", end_color="E0F5EC", fill_type="solid")
low_font = Font(name="Calibri", bold=True, size=10, color="1A7A4A")

enhance_row_fill = PatternFill(start_color="F0F3FF", end_color="F0F3FF", fill_type="solid")

body_font = Font(name="Calibri", size=10)
wrap_align = Alignment(vertical="top", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

headers = ["#", "Issue", "Type", "Priority", "Impact", "Description", "Area"]
col_widths = [6, 44, 16, 13, 12, 80, 22]

for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[chr(64 + col_idx)].width = w

ws.row_dimensions[1].height = 28

bugs = [
    [1, "Tab Resets / Gray Out & Performance Degradation", "Bug", "Critical", "High", "Core Navigation",
     "The site\u2019s tab frequently grays out and becomes unresponsive, triggered by: navigating too quickly between pages, or resizing panes (which can break/remove panes entirely). Additionally, tabs open for extended periods degrade in performance, becoming progressively slower. Each gray-out forces a full restart: new tab \u2192 sign in \u2192 navigate to area \u2192 search company \u2192 find product \u2192 reach target detail. During slow server periods this can repeat 6+ times in a row for a single field change. A task that should take seconds becomes minutes. This is the #1 productivity drain \u2014 it breaks focus, causes context loss (which client, which product, what was being adjusted), and the lengthy re-navigation path itself risks further gray-outs."],
    [2, "Summernote Save Workflow (Data Loss Risk)", "Bug / UX", "High", "High", "Code Editor",
     "Saving in the Summernote editor while in code view silently discards all changes. The user must remember to exit code view first \u2014 there is no warning or safeguard. Results in lost work and rework."],
    [3, "Product Save Workflow \u2014 Close & Reopen Per Tab", "Bug / UX", "High", "High", "Product Details",
     "Editing product details requires closing and reopening the product each time a different tab is modified. There is no way to save changes across multiple tabs in one session. Multi-field updates that should take a minute become a repetitive open-edit-close-reopen cycle."],
    [4, "FTP Definition Preview in Product Edit", "Enhancement", "High", "Medium", "Product Details / FTP",
     "No inline preview of the selected FTP definition when editing a product. To check the definition\u2019s structure (and determine whether additional FTP tab fields need to be filled in), the user must leave the product, navigate to the FTP definitions area, locate the definition, review it, then deep-path navigate back. An inline preview on selection would eliminate this detour entirely."],
    [5, "PDF Preview Missing in Composition", "Enhancement", "High", "Medium", "Composition",
     "The composition PDF view shows no preview, no filename, and no version info \u2014 just a download link with no context. Users must download the file to determine what version it is, making version comparison unnecessarily slow. Should display a preview (or at minimum filename + version label) and a clear download button."],
    [6, "Code Editor Improvements", "Enhancement", "High", "Medium", "Code Editor",
     "The embedded editor is missing standard features expected in any code editing environment: no multi-cursor (Ctrl/Cmd-D), no reliable in-place save (must click out of the editor to trigger save), and no basic code formatting or linting. These gaps slow daily work across every template edit."],
    [7, "Server Capacity / Sample Creation Timing", "Infrastructure", "Medium", "Medium", "Server / Data",
     "Sample creation runs on a fixed 5-minute interval rather than on-demand. Even a quick change requires waiting for the next cycle. This bottleneck may be a data server limitation. On-demand or near-immediate processing would tighten the edit-review loop significantly."],
    [8, "Symbol / Special Character Limitations (Wingdings)", "Bug / UX", "Medium", "Medium", "Code Editor / Summernote",
     "Special characters and symbols are restricted to the Wingdings font \u2014 likely due to Summernote, the print pipeline, or printer font constraints. Symbols that are clearly visible in the source Word document cannot be copy-pasted into the editor; instead, programmers must manually look up the corresponding Wingdings character code. A copy-paste workaround or a visual symbol picker mapping common symbols to their Wingdings equivalents would save significant time."],
    [9, "Error Transparency & Observability", "Enhancement", "Medium", "Medium", "Job Processing Screen",
     "Errors and failures across the platform are often silent or opaque. Suppressed files show no explanation, and error codes (when present) are not actionable. Users resort to manual investigation or guesswork. Clear, contextual error messages and accessible logs would reduce debugging time."],
    [10, "Dropdown Menus Overlap / Don\u2019t Auto-Close", "Bug", "Low", "Low", "Core UI",
     "Opening a new dropdown does not close the previously opened one, causing menus to overlap. Requires extra clicks to dismiss. Minor but persistent annoyance."],
    [11, "\u2018X\u2019 Button Misalignment on Company Filter", "Bug", "Low", "Low", "Composition Filter",
     "The close (\u2018X\u2019) button on the company filter in the composition filter menu does not align with the mouse click target. Hitbox is offset from the visible button."],
    [12, "Shift+Tab Navigation Broken on Dropdowns", "Bug", "Low", "Low", "UI / Accessibility",
     "Tabbing forward through form fields works normally, but Shift+Tab (backwards) gets stuck on dropdown components. If a user overshoots a field, they cannot tab back to it and must use the mouse. Only affects reverse tab order."],
    [13, "Page Refresh / Back-Forward Breaks Page Content", "Bug", "Low", "Low", "Core Navigation",
     "Refreshing the page or using the browser\u2019s back/forward buttons causes the page to break \u2014 only the blue title/header bar remains visible, with no content rendered below it and no buttons functional. This is distinct from the gray-out issue; the tab is still technically alive but completely unusable. Rare occurrence but unintuitive \u2014 standard browser actions shouldn\u2019t break the page."],
]

post_bugs = []

enhancements = [
    [14, "Two Separate Site Versions (SSO vs. Main)", "Bug", "Medium", "High", "Platform / Infrastructure",
     "Two separate versions of NcConnect are maintained under different URLs \u2014 one SSO, one non-SSO. Both must be updated independently, and bugs fixed in one version may not be fixed in the other, leading to inconsistent behavior. This should be consolidated into a single site (e.g., SSO-only with a settings toggle for auth method, or a unified codebase serving both). Maintaining two parallel versions doubles the update/testing burden and is a persistent source of divergent bugs."],
    [15, "Unified Workspace for Data Drops & File Access", "Enhancement", "Medium", "High", "Workflow / Files",
     "Viewing or dropping data currently requires leaving NcConnect and navigating deep server paths: client \u2192 product \u2192 date \u2192 production/dev \u2192 incoming/outgoing \u2192 target files. A typical data drop involves two of these navigations. Integrating file access and data dropping directly into NcConnect (a \u201cDrop Data\u201d component alongside one-click access to recent files for the current product) would eliminate these detours and keep the user in a single workflow."],
    [16, "Frontend Framework Evaluation (Blazor \u2192 React)", "Architecture", "Medium", "High", "Platform / Frontend",
     "The site is built on a Blazor-like C# framework. Evaluating a move to React or a similar modern JS framework could improve developer velocity, ecosystem support, and make the UX improvements in this list significantly easier to implement. Long-term initiative."],
]

priority_styles = {
    "Critical": (crit_fill, crit_font),
    "High": (high_fill, high_font),
    "Medium": (med_fill, med_font),
    "Low": (low_fill, low_font),
}

def write_row(ws, row_idx, row_data, is_enhancement=False):
    reordered = row_data[:5] + [row_data[6], row_data[5]]
    priority = reordered[3]
    p_fill, p_font = priority_styles.get(priority, (None, body_font))
    for col_idx, val in enumerate(reordered, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = body_font
        cell.alignment = wrap_align
        cell.border = thin_border
        if col_idx in (1, 3, 4, 5, 7):
            cell.alignment = center_align
        if col_idx == 4 and p_fill:
            cell.fill = p_fill
            cell.font = p_font
        if is_enhancement and col_idx not in (4,):
            cell.fill = enhance_row_fill
    ws.row_dimensions[row_idx].height = 60

row = 2
for item in bugs:
    write_row(ws, row, item)
    row += 1

for item in post_bugs:
    write_row(ws, row, item)
    row += 1

sep_row = row
for col_idx in range(1, 8):
    cell = ws.cell(row=sep_row, column=col_idx)
    if col_idx == 2:
        cell.value = "Additional items worth considering \u2014 these could help long-term and further improve speed / reduce friction"
        cell.font = Font(name="Calibri", bold=True, size=11, color="3355AA")
    cell.fill = PatternFill(start_color="D6DEFF", end_color="D6DEFF", fill_type="solid")
    cell.border = thin_border
    cell.alignment = Alignment(vertical="center", wrap_text=True)
ws.merge_cells(start_row=sep_row, start_column=2, end_row=sep_row, end_column=7)
ws.row_dimensions[sep_row].height = 32
row += 1

for item in enhancements:
    write_row(ws, row, item, is_enhancement=True)
    row += 1

ws.auto_filter.ref = f"A1:G{row - 1}"
ws.freeze_panes = "B2"
ws.sheet_properties.tabColor = "6C8CFF"

import shutil
repo_path = r"C:\Users\jhamrick\Desktop\NcFormatter\NcConnect_Bug_Tracker.xlsx"
desktop_path = r"C:\Users\jhamrick\Desktop\NcConnect_Bug_Tracker.xlsx"
wb.save(repo_path)
shutil.copy2(repo_path, desktop_path)
print(f"Saved to {repo_path} and {desktop_path}")
